import os

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, Alignment

from jira_git.csv import GitExportParser

class AbstractExcel(object):
    border = Border(left=Side(border_style='thin', color='FF000000'),
                    right=Side(border_style='thin', color='FF000000'),
                    top=Side(border_style='thin', color='FF000000'),
                    bottom=Side(border_style='thin', color='FF000000'),
                    diagonal=Side(border_style=None, color='FF000000'),
                    diagonal_direction=0,
                    outline=Side(border_style=None,
                                 color='FF000000'),
                    vertical=Side(border_style=None,
                                  color='FF000000'),
                    horizontal=Side(border_style=None,
                                    color='FF000000')
                    )

    alignment=Alignment(horizontal='general',
                            vertical='top',
                            text_rotation=0,
                            wrap_text=True,
                            shrink_to_fit=False,
                            indent=0)

    def __init__(self):
        self.pike_parser = GitExportParser()

    def _write_headers(self, sheet, row=1):
        column = 1
        for header in self.headers:
            sheet.cell(column=column, row=row, value=header[0])
            column += 1

    def setup_column_width(self, sheet):
        for header in self.headers:
            sheet.column_dimensions[header[1]].width = header[2]

class ExcelLineCounterReporter(AbstractExcel):

    def __init__(self):
        super(ExcelLineCounterReporter, self).__init__()
        self.headers = [['Base Folder', 'A', 15, 'base_folder'],
                        ['Filename', 'B', 15, 'filename'],
                        ['Full Path', 'C', 40, 'file'],
                        ['Extension', 'D', 8, 'extension'],
                        ['Count','E', 11, 'lines'],]

        self.sheet_name = 'Lines of Code'

    def write(self, output_filename, file_list, **kwargs):
        if os.path.exists(output_filename):
            wb = load_workbook(filename=output_filename)
            sheet = wb[self.sheet_name]
            row = sheet.get_highest_row()
        else:
            wb = Workbook()
            sheet = wb.create_sheet(title=self.sheet_name)
            self.setup_column_width(sheet)
            row = 1
            self._write_headers(sheet)

        for count_data in file_list:
            row += 1
            column = 1
            for header in self.headers:
                current_cell=sheet.cell(column=column, row=row, value=count_data[header[3]])
                current_cell.alignment = self.alignment
                current_cell.border = self.border
                column += 1

            current_cell.border = self.border
        wb.save(output_filename)
        return row - 1



class ExcelGitReporter(AbstractExcel):

    def __init__(self):
        super(ExcelGitReporter, self).__init__()

        self.headers = [['Commit Hash', 'A', 11, 'hash_id'],
                        ['Username','B', 11, 'username'],
                        ['Email','C', 25, 'email'],
                        ['Date', 'D', 19, 'date'],
                        ['Description', 'E', 50, 'description'],
                        ['Project', 'F', 19, 'project'],
                        ['Commit Type', 'G', 11, 'commit_type'],
                        ['Issue Number', 'H', 12, 'issue_number'],
                        ['Files changed', 'I', 12, 'files_changed'],
                        ['Insertions', 'J', 12, 'insertions'],
                        ['Deletions', 'K', 12, 'deletions'],
                        ['Repo', 'L', 20, 'repo_name'],
                        ['Stash Project', 'M', 20, 'stash_project'],
                        ['Branch', 'N', 12,'branch']]

    def write(self, output_filename, dictionary, **kwargs):
        if os.path.exists(output_filename):
            wb = load_workbook(filename=output_filename)
            sheet = wb['Commits']
            row = sheet.get_highest_row()
        else:
            wb = Workbook()
            sheet = wb.create_sheet(title='Commits')
            self.setup_column_width(sheet)
            row = 1
            self._write_headers(sheet)

        commits = self.pike_parser.parse_dictionary(dictionary, **kwargs)
        written_rows = 0
        for commit in commits:
            row += 1
            column = 1
            for header in self.headers:
                coordinate = header[1] + str(row)
                value = commit[header[3]]
                current_cell=sheet.cell(coordinate=coordinate, value=value)
                current_cell.alignment = self.alignment
                current_cell.border = self.border
                column += 1
            written_rows += 1
        wb.save(output_filename)
        return written_rows


class ExcelCommitImporter(AbstractExcel):
    '''
    Class to import list created with GitExportParser
    '''

    def __init__(self):
        super(ExcelCommitImporter, self).__init__()
        self.headers = [['Commit Hash', 'A', 11, 'hash_id'],
                        ['Username','B', 11, 'username'],
                        ['Date', 'C', 19, 'date'],
                        ['Description', 'D', 50, 'description'],
                        ['Project', 'E', 19, 'project'],
                        ['commit_type', 'F', 11, 'commit_type'],
                        ['issue_number','G', 12, 'issue_number'],
                        ['source_file', 'H', 12, 'base_filename']]

    def parse_folder(self, folder, output_filename, **kwargs):
        file_list= list()
        for root, dir, files in os.walk(folder):
            for file in files:
                filename = os.path.join(root, file)
                file_list.append(filename)
        return self.parse_multiple_files(file_list, output_filename, **kwargs)

    def parse_multiple_files(self, filenames, output_filename, **kwargs):
        wb = Workbook()
        sheet = wb.create_sheet(title='Commits')
        self.setup_column_width(sheet)
        row = 1
        written_rows = 0
        self._write_headers(sheet)
        for filename in filenames:
            commits = self.pike_parser.parse(filename, **kwargs)
            for commit in commits:
                row += 1
                column = 1
                for header in self.headers:
                    coordinate = header[1] + str(row)
                    value = commit[header[3]]
                    current_cell=sheet.cell(coordinate=coordinate, value=value)
                    current_cell.alignment = self.alignment
                    current_cell.border = self.border
                    column += 1
                written_rows += 1
            # for commit in commits:
            #     row += 1
            #     column = 1
            #     current_cell=sheet.cell(column=column, row=row, value=commit[0])
            #     current_cell.alignment = self.alignment
            #     current_cell.border = self.border
            #
            #     column += 1
            #     current_cell=sheet.cell(column=column, row=row, value=commit[1])
            #     current_cell.alignment = self.alignment
            #     current_cell.border = self.border
            #
            #     column += 1
            #     current_cell=sheet.cell(column=column, row=row, value=commit[2])
            #     current_cell.alignment = self.alignment
            #     current_cell.border = self.border
            #
            #     column += 1
            #     current_cell=sheet.cell(column=column, row=row, value=commit[3])
            #     current_cell.alignment = self.alignment
            #     current_cell.border = self.border
            #
            #     column += 1
            #     current_cell=sheet.cell(column=column, row=row, value=commit[4])
            #     current_cell.alignment = self.alignment
            #     current_cell.border = self.border
            #
            #     column += 1
            #     current_cell=sheet.cell(column=column, row=row, value=commit[5])
            #     current_cell.alignment = self.alignment
            #     current_cell.border = self.border
            #
            #     column += 1
            #     current_cell=sheet.cell(column=column, row=row, value=commit[6])
            #     current_cell.alignment = self.alignment
            #     current_cell.border = self.border
            #
            #     column += 1
            #     current_cell=sheet.cell(column=column, row=row, value=commit[7])
            #     current_cell.alignment = self.alignment
            #     current_cell.border = self.border

        wb.save(output_filename)
        return written_rows

