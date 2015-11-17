from datetime import datetime
import os
from django.core.management import BaseCommand
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

from common.utils import filename_with_datetime
from employee.models import CompanyGroup
from homeworkpal_project.settings.base import TEST_OUTPUT_PATH
from project_admin.models import IndividualGoal
from django.utils import timezone
__author__ = 'luiscberrocal'


class Command(BaseCommand):
    '''
    To export to excel the individual goals of a group
    python manage.py export_goals TINO-NS
    '''

    border = Border(left=Side(border_style='thin', color='FF000000'),
                    right=Side(border_style='thin', color='FF000000'),
                    top=Side(border_style='thin', color='FF000000'),
                    bottom=Side(border_style='thin', color='FF000000'),
                    diagonal=Side(border_style=None, color='FF000000'),
                    diagonal_direction=0,
                    outline=Side(border_style=None,
                                 color='FF000000'),
                    vertical=Side(border_style=None,
                                  color='FF000000'),
                    horizontal=Side(border_style=None,
                                    color='FF000000')
                    )

    alignment=Alignment(horizontal='general',
                            vertical='top',
                            text_rotation=0,
                            wrap_text=True,
                            shrink_to_fit=False,
                            indent=0)

    def add_arguments(self, parser):
        parser.add_argument('group')

    def handle(self, *args, **options):
        group = CompanyGroup.objects.get(name=options['group'])
        employees = group.members()
        filename = filename_with_datetime(TEST_OUTPUT_PATH, '%s.xlsx' % (options['group']))
        #os.path.join(TEST_OUTPUT_PATH, '%s_%s.xlsx' % (options['group'], timezone.now().strftime('%Y%m%d_%H%M')))
        wb = Workbook()
        goals = IndividualGoal.objects.filter(employee__in=employees).order_by('employee', 'priority')
        pos = 1
        current_username = None
        headers = ['No', 'Meta', 'Descripción', 'Totalmente Satifactorio', 'Peso', 'Entregables']
        for goal in goals:
            if goal.employee.user.username != current_username:
                current_username = goal.employee.user.username
                pos = 1
                sheet = wb.create_sheet(title=current_username)
                sheet['A1'] = '%s  IP: %s' % (goal.employee, goal.employee.company_id)
                self._title_font(sheet['A1'])
                col = 1
                for title in headers:
                    cell = sheet.cell(column=col, row=3, value=title)
                    cell.border = self.border
                    cell.alignment = self.alignment
                    self._format_header(cell)
                    col += 1
                row = 4
            ## No Proyecto A
            col = 1
            cell = sheet.cell(column=col, row=row, value=pos)
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ## Proyecto B
            col += 1
            cell = sheet.cell(column=col, row=row, value=str(goal.name))
            cell.border = self.border
            self._wrap_cell(cell)
            ## Description C
            col += 1
            cell = sheet.cell(column=col, row=row, value=str(goal.description))
            self._wrap_cell(cell)
            cell.border = self.border
            ## Satifactory D
            col += 1
            cell = sheet.cell(column=col, row=row, value=goal.expectations)
            self._wrap_cell(cell)
            cell.border = self.border
            ## Weight E
            col += 1
            cell = sheet.cell(column=col, row=row, value=goal.weight)
            cell.number_format = '0%'
            cell.alignment = Alignment(horizontal='general', vertical='top')
            cell.border = self.border
            ## Deliverbales F
            col += 1
            deliverables = ''
            if goal.project:
                for deliverable in goal.project.deliverables.all():
                    deliverables += deliverable.name
                    deliverables += '\n'
            else:
                deliverables = 'No Aplica'
            cell = sheet.cell(column=col, row=row, value=deliverables)
            self._wrap_cell(cell)
            cell.border = self.border

            row += 1

            self._set_column_widths(sheet)
            self._page_setup(sheet)
            print('%d %-50s %s' % (pos, goal.name, goal.employee))
            pos += 1

        wb.save(filename)
        self.stdout.write('Wrote %s'  % filename)

    def _wrap_cell(self, cell):
        cell.alignment = self.alignment

    def _page_setup(self, sheet):
        sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
        sheet.page_setup.fitToWidth = 1
        sheet.header_footer.center_header.text = 'Metas Individuales de Desempeño\nAño Fiscal 2015'
        sheet.header_footer.center_header.font_size = 20
        sheet.header_footer.center_header.font_name = "Tahoma,Bold"
        sheet.header_footer.right_header.text = '&[Date]'
        sheet.header_footer.right_footer.text = '&[Page] de &[Pages]'
        sheet.header_footer.left_footer.text= '______________________________\nFirma'

    def _set_column_widths(self, sheet):
        sheet.column_dimensions["A"].width = 5.0
        sheet.column_dimensions["B"].width = 30.0
        sheet.column_dimensions["C"].width = 50.0
        sheet.column_dimensions["D"].width = 37.0
        sheet.column_dimensions["E"].width = 6.0
        sheet.column_dimensions["F"].width = 37.0

    def _title_font(self, cell):
        font = font = Font(name='Calibri', size=18, bold=True, italic=False,
                           vertAlign=None, underline='none', strike=False, color='FF000000')

        cell.font = font

    def _format_header(self, cell):
        font = Font(name='Calibri', size=12, bold=True, italic=False,
                    vertAlign=None, underline='none', strike=False, color='FF000000')
        fill = PatternFill(fill_type='solid', start_color='CBE800', end_color='CBE800')
        cell.font = font
        cell.fill = fill





