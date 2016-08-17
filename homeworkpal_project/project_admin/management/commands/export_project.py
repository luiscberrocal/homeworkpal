from datetime import datetime
import os
from django.core.management import BaseCommand
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from common.utils import filename_with_datetime
from homeworkpal_project.settings.base import TEST_DATA_PATH
from project_admin.models import Project

__author__ = 'lberrocal'


class Command(BaseCommand):
    def add_arguments(self, parser):
        parser.add_argument('project_id', nargs='?')
        parser.add_argument('--all',
                            action='store_true',
                            dest='process_all',
                            help='Export all projects')

        parser.add_argument('--dashboard',
                            action='store_true',
                            dest='export_dashboard',
                            help='Export dashboard projects')

    def handle(self, *args, **options):
        if options['process_all']:
            self.export_all_projects()
        elif ['export_dashboard']:
            self.export_on_dashboard()
        else:
            self.export_single_project(options)

    def export_all_projects(self):
        row = 1
        wb = Workbook()
        sheet = wb.active
        output_filename = filename_with_datetime(TEST_DATA_PATH, 'tino-ns-projects.xlsx')
        for project in Project.objects.all():
            sheet['A%d' % row] = project.short_name
            sheet['B%d' % row] = project.description
            sheet['C%d' % row] = project.actual_start_date
            sheet['D%d' % row] = project.actual_end_date
            row += 1
        wb.save(output_filename)
        self.stdout.write('Wrote: %s' % (output_filename))

    def export_on_dashboard(self):
        row = 1
        wb = Workbook()
        sheet = wb.active
        output_filename = filename_with_datetime(TEST_DATA_PATH, 'tino-ns-db-projects.xlsx')
        for project in Project.objects.filter(on_dashboard=True):
            sheet['A%d' % row] = project.short_name
            sheet['B%d' % row] = project.description
            sheet['C%d' % row] = project.actual_start_date
            sheet['D%d' % row] = project.actual_end_date
            row += 1
        wb.save(output_filename)
        self.stdout.write('Wrote: %s' % (output_filename))

    def export_single_project(self, options):
        project = Project.objects.get(pk=int(options['project_id']))
        template_filename = os.path.join(TEST_DATA_PATH, '1680_v2.xlsx')
        now = timezone.localtime(timezone.now())
        output_filename = os.path.join(TEST_DATA_PATH,
                                       '%s_%s.xlsx' % (project.slug.replace('-', '_'), now.strftime('%Y%m%d_%H%M')))
        wb = load_workbook(template_filename)
        sheet = wb.active
        sheet['D2'] = project.short_name
        sheet['I3'] = datetime.today()
        sheet['A9'] = project.description
        row = 12
        for corporate_goal_assignment in project.corporate_goals.all()[:4]:
            sheet['A%d' % row] = corporate_goal_assignment.corporate_goal.number
            sheet['B%d' % row] = corporate_goal_assignment.corporate_goal.description
            row += 1
        row = 20
        for deliverable in project.deliverables.all()[:5]:
            sheet['B%d' % row] = deliverable.name
            sheet['G%d' % row] = deliverable.description
            row += 1
        row = 27
        for stakeholder in project.stakeholders.all()[:5]:
            sheet['B%d' % row] = str(stakeholder.employee)
            # sheet['G%d'%row] = deliverable.description
            row += 1
        row = 42
        for risk in project.risks.all()[:5]:
            sheet['B%d' % row] = risk.risk_type
            sheet['C%d' % row] = risk.description
            row += 1

        row = 49
        sheet['A%d' % row] = str(project.planned_start_date)
        sheet['H%d' % row] = str(project.planned_end_date)
        wb.save(output_filename)
        self.stdout.write('Wrote %s' % output_filename)
