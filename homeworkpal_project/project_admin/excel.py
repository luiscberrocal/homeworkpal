from openpyxl import Workbook

__author__ = 'LBerrocal'


class ProjectAdminExcel(object):

    def export_supports(self, filename, supports):
        wb = Workbook()
        sheet = wb.create_sheet(title='Supports')
        row = 1
        headers = ['Project', 'Group', 'Date', 'Description']
        column = 1
        for header in headers:
            sheet.cell(column=column, row=row, value=header)
            column += 1
        for support in supports:
            row += 1
            column = 1
            sheet.cell(column=column, row=row, value=support.project.short_name)
            column += 1
            sheet.cell(column=column, row=row, value=support.company_group.name)
            column += 1
            sheet.cell(column=column, row=row, value=support.required_date)
            column += 1
            sheet.cell(column=column, row=row, value=str(support.description))
        wb.save(filename)

