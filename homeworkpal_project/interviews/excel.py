from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill

__author__ = 'LBerrocal'

class ExcelCertificateResults(object):

    border = Border(left=Side(border_style='thin', color='FF000000'),
                    right=Side(border_style='thin', color='FF000000'),
                    top=Side(border_style='thin', color='FF000000'),
                    bottom=Side(border_style='thin', color='FF000000'),
                    diagonal=Side(border_style=None, color='FF000000'),
                    diagonal_direction=0,
                    outline=Side(border_style=None,
                                 color='FF000000'),
                    vertical=Side(border_style=None,
                                  color='FF000000'),
                    horizontal=Side(border_style=None,
                                    color='FF000000')
                    )

    alignment=Alignment(horizontal='general',
                            vertical='top',
                            text_rotation=0,
                            wrap_text=True,
                            shrink_to_fit=False,
                            indent=0)

    def export(self, filename, certificate):
        wb = Workbook()
        sheet = wb.create_sheet(title='Resultados')
        self.setup_column_width(sheet)

        row = 1
        headers = ['Reultado de la Selección', 'Cédula', 'Nombre', 'Entrevistadores', 'Comentarios']
        column = 1
        for header in headers:
            current_cell = sheet.cell(column=column, row=row, value=header)
            current_cell.alignment = self.alignment
            current_cell.border = self.border
            self._format_header(current_cell)
            column += 1

        for candidate in certificate.candidates.all():
            row += 1
            column = 1
            current_cell = sheet.cell(column=column, row=row, value=str(candidate.certificate_result))
            current_cell.alignment = self.alignment
            current_cell.border = self.border
            column += 1
            current_cell = sheet.cell(column=column, row=row, value=str(candidate.candidate.national_id))
            current_cell.alignment = self.alignment
            current_cell.border = self.border
            column += 1
            current_cell = sheet.cell(column=column, row=row, value=str(candidate.candidate))
            current_cell.alignment = self.alignment
            current_cell.border = self.border
            column += 1
            current_cell = sheet.cell(column=column, row=row, value='Luis Berrocal')
            current_cell.alignment = self.alignment
            current_cell.border = self.border
            column += 1
            current_cell = sheet.cell(column=column, row=row, value=candidate.explanation)
            current_cell.alignment = self.alignment
            current_cell.border = self.border

        self._page_setup(sheet, certificate.number, certificate.grade)
        wb.save(filename)

    def setup_column_width(self, sheet):
        sheet.column_dimensions['A'].width = 9
        sheet.column_dimensions['B'].width = 12
        sheet.column_dimensions['C'].width = 25
        sheet.column_dimensions['D'].width = 15
        sheet.column_dimensions['E'].width = 60

    def _page_setup(self, sheet, certificate_number, grade):
        sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
        sheet.page_setup.fitToWidth = 1
        sheet.header_footer.center_header.text = 'Certificado %s %s' % (certificate_number, grade)
        sheet.header_footer.center_header.font_size = 20
        sheet.header_footer.center_header.font_name = "Tahoma,Bold"
        sheet.header_footer.right_header.text = '&[Date]'
        sheet.header_footer.right_footer.text = '&[Page] de &[Pages]'
        #sheet.header_footer.left_footer.text= '______________________________\nFirma'


    def _format_header(self, cell):
        font = Font(name='Calibri', size=12, bold=True, italic=False,
                    vertAlign=None, underline='none', strike=False, color='FF000000')
        fill = PatternFill(fill_type='solid', start_color='CBE800', end_color='CBE800')
        cell.font = font
        cell.fill = fill