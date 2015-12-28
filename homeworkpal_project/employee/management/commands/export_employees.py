from django.core.management import BaseCommand
from openpyxl import Workbook

from common.utils import filename_with_datetime
from employee.models import Employee
from homeworkpal_project.settings.base import TEST_OUTPUT_PATH


__author__ = 'lberrocal'


class Command(BaseCommand):
    def add_arguments(self, parser):
        parser.add_argument('company_group')

    def handle(self, *args, **options):
        from maximo.models import  MaximoTimeRegister
        company_group = options['company_group']
        self.stdout.write('company_group: %s' % company_group)
        filename = filename_with_datetime(TEST_OUTPUT_PATH, 'Export_Employees_%s.xlsx' % (company_group))
        employees = Employee.objects.from_group(company_group)
        wb = Workbook()
        sheet = wb.create_sheet(title='Employees')
        row = 1
        headers = ['Company Id', 'Username', 'Grade', 'Position Owner', 'Position Number',  'Position Tenure','Pay Rate']
        column = 1
        for header in headers:
            sheet.cell(column=column, row=row, value=header)
            column += 1
        for employee in employees:
            row += 1
            column = 1
            sheet.cell(column=column, row=row, value=employee.company_id)
            column += 1
            sheet.cell(column=column, row=row, value=employee.user.username)
            column += 1
            sheet.cell(column=column, row=row, value=employee.position.grade)
            column += 1
            if employee.position.owner:
                sheet.cell(column=column, row=row, value=employee.position.owner.user.username)
            else:
                sheet.cell(column=column, row=row, value='NA')
            column += 1
            sheet.cell(column=column, row=row, value=employee.position.number)
            column += 1
            sheet.cell(column=column, row=row, value=employee.position.type)
            ticket = MaximoTimeRegister.objects.filter(employee=employee).order_by('date')[1:]
            column += 1
            sheet.cell(column=column, row=row, value=ticket[0].pay_rate)
        wb.save(filename)
        self.stdout.write('File export %s' % filename)

        