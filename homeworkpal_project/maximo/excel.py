import csv
from datetime import datetime, time
from decimal import Decimal
from openpyxl import load_workbook, Workbook
from employee.models import Employee
from .models import MaximoTicket, MaximoTimeRegister

import logging

logger = logging.getLogger(__name__)
__author__ = 'lberrocal'


def row_to_dictionary(excel_row, mappings):
    data = dict()
    for attribute, position in mappings.items():
        data[attribute] = excel_row[position].value
    return data


def parse_hours(str_hours):
    parts = str_hours.split(':')
    return Decimal(parts[0]) + Decimal(parts[1]) / 60


def parse_datetime_hours(hours):
    return Decimal(hours.hour + hours.minute / 60.0)


def decimal_to_time(decimal_hours):
    hour = int(decimal_hours)
    minute = int((decimal_hours - int(decimal_hours)) * Decimal(60.0))
    return time(hour, minute, 0)


class AbstractMaximoData(object):
    '''
    The report TINO-NS-FY16 has the following columns
        acp_empnum	0
        acp_hours	1
        acp_pagorelevo	2
        acp_timingdate	3
        enterby	4
        laborcode	5
        memo	6
        payrate	7
        refwo	8
        regularhrs	9
        skilllevel	10
        ticketclass	11
        ticketid	12
        total_rec	13
        params_where	14

    '''
    LOAD_TICKETS = 'LOAD_TICKETS'
    LOAD_TIME = 'LOAD_TIME'
    LOAD_ALL = 'LOAD_ALL'

    def __init__(self, stdout=None, **kwargs):
        self.ticket_mappings = {'ticket_type': 0, 'number': 1, 'name': 2}
        self.time_register_mappings = {'company_id': 0, #acp_num on the report
                                       'regular_hours': 1, #acp_hours on the report
                                       'date': 3, #acp_timpingdate on the report
                                       'username': 5, #laborcode on the report
                                       'pay_rate': 7,
                                       'wo_number': 8,
                                       'ticket_type': 11,
                                       'ticket_number': 12,
                                       'description': 6}
        self.stdout = stdout


    def write(self, msg):
        if self.stdout:
            self.stdout.write(msg)

    def _get_maximo_ticket_info(self, row):
        valid_ticket_types =  [MaximoTicket.MAXIMO_WORKORDER, MaximoTicket.MAXIMO_INCIDENT, MaximoTicket.MAXIMO_SR]
        ticket_type = row[self.time_register_mappings['ticket_type']]
        if not isinstance(ticket_type, str):
            ticket_type = ticket_type.value
        if ticket_type is None:
            ticket_type = MaximoTicket.MAXIMO_WORKORDER
        assert ticket_type is not None, 'Could not find ticket type'
        ticket_type = ticket_type.strip()
        assert ticket_type in valid_ticket_types, 'Invalid ticket type: "%s"' % ticket_type
        if ticket_type in valid_ticket_types[:1]:
            number = row[self.time_register_mappings['wo_number']]
        else:
            number = row[self.time_register_mappings['ticket_number']]
        if not isinstance(number, str):
            number = number.value
        return ticket_type, number


class MaximoCSVData(AbstractMaximoData):

    def __init__(self, stdout=None, **kwargs):
        super(MaximoCSVData, self).__init__(stdout=stdout, **kwargs)
        self.delimiter = kwargs.get('delimiter', ',')

    def _parse_date(self, str_date):
        return datetime.strptime(str_date, '%b %d, %Y').date()

    def load_time_registers(self, filename):
        time_results = {'rows_parsed': 0,
                        'created': 0,
                        'duplicates': 0,
                        'sheet': 'NA',
                        'errors': list()}
        row_num = 1
        created_count = 0
        updated = 0
        duplicate_count = 0
        errors = list()
        with open(filename, 'r', encoding='utf-8') as csv_file:
            csv_reader = csv.reader(csv_file, delimiter=self.delimiter)
            next(csv_reader, None)
            for row in csv_reader:
                attributes = dict()
                company_id = row[self.time_register_mappings['company_id']]
                try:
                    attributes['employee'] = Employee.objects.get(company_id=company_id)
                    attributes['date'] = self._parse_date(row[self.time_register_mappings['date']])
                    regular_hours = parse_hours(row[self.time_register_mappings['regular_hours']])
                    if regular_hours > 8.0:
                        raise ValueError(
                            'Regular hours cannot exceed 8 hours. Your are trying to add %.1f hours' % regular_hours)
                    register_summary = MaximoTimeRegister.objects.get_employee_total_regular_hours(**attributes)
                    total_regular_hours = 0
                    if register_summary['total_regular_hours'] is not None:
                        total_regular_hours = register_summary['total_regular_hours']
                    if total_regular_hours + regular_hours <= 8.0:
                        attributes['pay_rate'] = Decimal(row[self.time_register_mappings['pay_rate']])
                        ticket_type, number = self._get_maximo_ticket_info(row)
                        attributes['ticket'] = MaximoTicket.objects.get(ticket_type=ticket_type, number=number)
                        attributes['regular_hours'] = regular_hours
                        attributes['defaults'] = {'description': row[self.time_register_mappings['description']]}
                        register, created = MaximoTimeRegister.objects.get_or_create(**attributes)
                        if created:
                            created_count += 1
                        else:
                            msg = 'Data on row %d for employee %s  ' \
                                  'seems to be duplicated for record %d' % (row_num, attributes['employee'],
                                                                            register.pk)
                            logger.warn(msg)
                            error = {'row_num': row_num,
                                     'type': 'Possible duplicate',
                                     'message': msg}
                            errors.append(error)
                            duplicate_count += 1
                    else:
                        msg = 'Data on row %d for employee %s exceeds ' \
                              'the maximum regular hour. It would end up having %.1f hours' % (row_num,
                                                                                               attributes['employee'],
                                                                                               total_regular_hours + regular_hours)
                        logger.warn(msg)
                        error = {'row_num': row_num,
                                 'type': 'Exceed maximum 8 regular hours',
                                 'message': msg}
                        errors.append(error)
                        duplicate_count += 1
                except Employee.DoesNotExist:
                    username = row[self.time_register_mappings['username']]
                    msg = 'Employee with id %s and username %s ' \
                          'on row %d does not exist time registe was not loaded' % (company_id, username, row_num)
                    logger.error(msg)
                    error = {'row_num': row_num,
                             'type': 'Employee does not exist',
                             'message': msg}
                    errors.append(error)
                    #raise ValueError()
                except MaximoTicket.DoesNotExist:
                    msg = '%s with number %s on line %d does not exist' % (ticket_type, number, row_num)
                    logger.error(msg)
                    error = {'row_num': row_num,
                             'type': 'Ticket does not exist',
                             'message': msg}
                    errors.append(error)
                except TypeError as te:
                    msg = 'Unexpected error %s on row %d' % (te, row_num)
                    logger.error(msg)
                    error = {'row_num': row_num,
                             'type': 'Unexeptected Type Error',
                             'message': msg}
                    errors.append(error)
                except ValueError as ve:
                    msg = '%s on row %d' % (ve, row_num)
                    logger.error(msg)
                    error = {'row_num': row_num,
                             'type': 'Value Error',
                             'message': msg}
                    errors.append(error)
                except AssertionError as e:
                    msg = '%s on row %d' % (e, row_num)
                    logger.error(msg)
                    error = {'row_num': row_num,
                             'type': 'Assertion Error',
                             'message': msg}
                    errors.append(error)

                row_num += 1
        time_results['rows_parsed'] = row_num - 1
        time_results['created'] = created_count
        time_results['duplicates'] = duplicate_count
        time_results['errors'] = errors
        return time_results




class MaximoExcelData(AbstractMaximoData):
    '''
    The loading of Times is based on an export of a report name TINO-NS-FY16.
    The columns are:
    0  acp_empnum       company_id
    1  acp_hours        regular_hours
    2  acp_pagorelevo
    3  acp_timingdate   date
    4  enterby
    5  laborcode        username
    6  memo             description
    7  payrate          pay_rate
    8  refwo            wo_number
    9  regularhrs
    10 skilllevel
    11 ticketclass      ticket_type
    12 ticketid         ticket_number
    '''


    def __init__(self, stdout=None):
        super(MaximoExcelData, self).__init__(stdout=stdout)
        self.ticket_sheet = 'Maximo Tickets'
        self.time_sheet = 'Time'

    def load(self, filename, action=AbstractMaximoData.LOAD_ALL, allow_update=False, **kwargs):
        wb = load_workbook(filename=filename, data_only=True)
        ticket_results = dict()
        time_results = dict()
        if action == self.LOAD_TICKETS:
            ticket_results = self.load_tickets(wb, allow_update=allow_update, **kwargs)
        elif action == self.LOAD_TIME:
            time_results = self.load_time_registers(wb, allow_update=allow_update, **kwargs)
        elif action is self.LOAD_ALL:
            ticket_results = self.load_tickets(wb, allow_update=allow_update, **kwargs)
            time_results = self.load_time_registers(wb, allow_update=allow_update, **kwargs)
        else:
            raise ValueError('"%s" is an invalid action for load' % action)

        return {'ticket_results': ticket_results,
                'time_results': time_results}

    def save_tickets(self, filename, tickets):
        wb = Workbook()
        sheet = wb.create_sheet(title=self.ticket_sheet)
        row = 1
        for v, column in self.ticket_mappings.items():
            sheet.cell(column=column + 1, row=row, value=v.upper())
        row += 1
        for ticket in tickets:
            for v, column, in self.ticket_mappings.items():
                sheet.cell(column=column + 1, row=row, value=getattr(ticket, v))
            row += 1

        wb.save(filename)

    def export_time_registers(self, filename, registers):
        wb = Workbook()
        sheet = wb.create_sheet(title=self.time_sheet)
        row = 1
        headers = ['Company Id', 'Username', 'Date', 'Hours', 'Pay Rate',
                   'Ticket Type', 'Ticket Number', 'Ticket Name', 'Memo', 'Project', 'Project Source']
        column = 1
        for header in headers:
            sheet.cell(column=column, row=row, value=header)
            column += 1
        for register in registers:
            row += 1
            column = 1
            sheet.cell(column=column, row=row, value=register.employee.company_id)
            column += 1
            sheet.cell(column=column, row=row, value=register.employee.user.username)
            column += 1
            sheet.cell(column=column, row=row, value=register.date)
            column += 1
            sheet.cell(column=column, row=row, value=register.regular_hours)
            column += 1
            sheet.cell(column=column, row=row, value=register.pay_rate)
            column += 1
            sheet.cell(column=column, row=row, value=register.ticket.ticket_type)
            column += 1
            sheet.cell(column=column, row=row, value=register.ticket.number)
            column += 1
            sheet.cell(column=column, row=row, value=register.ticket.name)
            column += 1
            sheet.cell(column=column, row=row, value=register.description)
            column += 1
            if register.ticket.project:
                project_name = register.ticket.project.short_name
            else:
                project_name=''
            sheet.cell(column=column, row=row, value=project_name)
            column += 1
            sheet.cell(column=column, row=row, value='NA')
        wb.save(filename)

    def save_time_registers(self, filename, registers):
        """
        Saves a queryset of MaximoTimeRegister objects to an excel format that matches the load file format. The load
        file forma is explained at the class level documentacions
        :param filename: Excel filename to save the MaximoTimeRegister
        :param registers: QuerySet of MaximoTimeRegister
        :return: None
        """
        wb = Workbook()
        sheet = wb.create_sheet(title=self.time_sheet)
        row = 1
        for v, column in self.time_register_mappings.items():
            sheet.cell(column=column + 1, row=row, value=v.upper())
        row += 1
        for register in registers:
            col = self.time_register_mappings['company_id'] + 1
            sheet.cell(column=col, row=row, value=register.employee.company_id)
            col = self.time_register_mappings['regular_hours'] + 1
            hours = decimal_to_time(register.regular_hours)
            sheet.cell(column=col, row=row, value=hours)
            col = self.time_register_mappings['date'] + 1
            sheet.cell(column=col, row=row, value=register.date)
            col = self.time_register_mappings['username'] + 1
            sheet.cell(column=col, row=row, value=register.employee.user.username)
            col = self.time_register_mappings['pay_rate'] + 1
            sheet.cell(column=col, row=row, value=register.pay_rate)
            col = self.time_register_mappings['description'] + 1
            sheet.cell(column=col, row=row, value=register.description)
            if register.ticket.ticket_type == MaximoTicket.MAXIMO_WORKORDER:
                col = self.time_register_mappings['wo_number'] + 1
                sheet.cell(column=col, row=row, value=register.ticket.number)
            if register.ticket.ticket_type != MaximoTicket.MAXIMO_WORKORDER:
                col = self.time_register_mappings['ticket_type'] + 1
                sheet.cell(column=col, row=row, value=register.ticket.ticket_type)
                col = self.time_register_mappings['ticket_number'] + 1
                sheet.cell(column=col, row=row, value=register.ticket.number)
            row += 1
        wb.save(filename)

    def load_time_registers(self, wb, allow_update=False, **kwargs):
        sheet_name = kwargs.get('Time', self.time_sheet)
        time_sheet = wb[sheet_name]
        time_results = {'rows_parsed': 0,
                        'created': 0,
                        'duplicates': 0,
                        'sheet': sheet_name,
                        'errors': list()}
        row_num = 1
        created_count = 0
        updated = 0
        duplicate_count = 0
        errors = list()
        for row in time_sheet.rows:
            if row_num > 1:
                attributes = dict()
                company_id = row[self.time_register_mappings['company_id']].value
                try:
                    attributes['employee'] = Employee.objects.get(company_id=company_id)
                    attributes['date'] = row[self.time_register_mappings['date']].value
                    regular_hours = parse_datetime_hours(row[self.time_register_mappings['regular_hours']].value)
                    if regular_hours > 8.0:
                        raise ValueError(
                            'Regular hours cannot exceed 8 hours. Your are trying to add %.1f hours' % regular_hours)
                    register_summary = MaximoTimeRegister.objects.get_employee_total_regular_hours(**attributes)
                    total_regular_hours = 0
                    if register_summary['total_regular_hours'] is not None:
                        total_regular_hours = register_summary['total_regular_hours']
                    if total_regular_hours + regular_hours <= 8.0:
                        attributes['pay_rate'] = Decimal(row[self.time_register_mappings['pay_rate']].value)
                        ticket_type, number = self._get_maximo_ticket_info(row)
                        attributes['ticket'] = MaximoTicket.objects.get(ticket_type=ticket_type, number=number)
                        attributes['regular_hours'] = regular_hours
                        attributes['defaults'] = {'description': row[self.time_register_mappings['description']].value}
                        register, created = MaximoTimeRegister.objects.get_or_create(**attributes)
                        if created:
                            created_count += 1
                        else:
                            msg = 'Data on row %d for employee %s  ' \
                                  'seems to be duplicated for record %d' % (row_num, attributes['employee'],
                                                                            register.pk)
                            logger.warn(msg)
                            error = {'row_num': row_num,
                                     'type': 'Possible duplicate',
                                     'message': msg}
                            errors.append(error)
                            duplicate_count += 1
                    else:
                        msg = 'Data on row %d for employee %s exceeds ' \
                              'the maximum regular hour. It would end up having %.1f hours' % (row_num,
                                                                                               attributes['employee'],
                                                                                               total_regular_hours + regular_hours)
                        logger.warn(msg)
                        error = {'row_num': row_num,
                                 'type': 'Exceed maximum 8 regular hours',
                                 'message': msg}
                        errors.append(error)
                        duplicate_count += 1
                except Employee.DoesNotExist:
                    username = row[self.time_register_mappings['username']].value
                    msg = 'Employee with id %s and username %s ' \
                          'on row %d does not exist time registe was not loaded' % (company_id, username, row_num)
                    logger.warn(msg)
                    error = {'row_num': row_num,
                             'type': 'Employee does not exist',
                             'message': msg}
                    errors.append(error)
                except MaximoTicket.DoesNotExist:
                    msg = '%s with number %s on line %d does not exist' % (ticket_type, number, row_num)
                    logger.warn(msg)
                    error = {'row_num': row_num,
                             'type': 'Ticket does not exist',
                             'message': msg}
                    errors.append(error)
                except TypeError as te:
                    msg = 'Unexpected error %s on row %d' % (te, row_num)
                    logger.error(msg)
                    error = {'row_num': row_num,
                             'type': 'Unexeptected Type Error',
                             'message': msg}
                    errors.append(error)
                except ValueError as ve:
                    msg = '%s on row %d' % (ve, row_num)
                    logger.error(msg)
                    error = {'row_num': row_num,
                             'type': 'Value Error',
                             'message': msg}
                    errors.append(error)
            row_num += 1
        time_results['rows_parsed'] = row_num - 2
        time_results['created'] = created_count
        time_results['duplicates'] = duplicate_count
        time_results['sheet'] = sheet_name
        time_results['errors'] = errors
        return time_results

    def load_tickets(self, wb, allow_update=False, **kwargs):
        sheet_name = kwargs.get('ticket_sheet', self.ticket_sheet)
        ticket_sheet = wb[sheet_name]
        results = {'rows_parsed': 0,
                   'created': 0,
                   'updated': 0,
                   'sheet': sheet_name}
        row_num = 1
        created_count = 0
        updated = 0
        for row in ticket_sheet.rows:
            if row_num > 1:
                data_dictionary = row_to_dictionary(row, self.ticket_mappings)
                obj, created = MaximoTicket.objects.get_or_create(ticket_type=data_dictionary['ticket_type'],
                                                                  number=data_dictionary['number'],
                                                                  defaults=data_dictionary)
                if created:
                    self.write('%d Created Maximo ticket %s' % (row_num - 1, obj))
                    logger.debug('%d Created Maximo ticket %s' % (row_num - 1, obj))
                    created_count += 1
                    # logger.debug('--- %d tickets created' % created)
                elif allow_update:
                    self.write('%d Update Maximo ticket %s' % (row_num - 1, obj))
                    logger.debug('%d Update Maximo ticket %s' % (row_num - 1, obj))
                    updated += 1
                else:
                    logger.debug('%d Existed Maximo ticket %s' % (row_num - 1, obj))
            row_num += 1
        # logger.debug('%d tickets created' % created)
        results = {'rows_parsed': row_num - 2,
                   'created': created_count,
                   'updated': updated,
                   'sheet': sheet_name}
        return results
